"""
Parse every participant workbook in participants/ into a normalised
predictions JSON in data/predictions/.

Each participant's filled-in template gives us:
  * predicted score for every group match            (cols E/F/G/H)
  * predicted group winner & runner-up per group      (place 1 / place 2)
  * predicted teams reaching each knockout round       (bracket regions)
  * predicted World Champion                           (cell CK68)
  * predicted top scorer (if they typed one, else from data/topscorers.csv)

Run:
    python src/parse_predictions.py
"""
from __future__ import annotations

import csv
import glob
import json
import os

import openpyxl
from openpyxl.utils import column_index_from_string as ci

import config as C


def _cell(ws, col_letter, row):
    return ws.cell(row=row, column=ci(col_letter)).value


def _participant_name(path):
    base = os.path.splitext(os.path.basename(path))[0]
    # "edvard_tipping" -> "Edvard"
    name = base.replace("_tipping", "").replace("_", " ").strip()
    return name.title() if name else base


def _load_topscorer_csv():
    """name (lowercased) -> top scorer pick, from data/topscorers.csv."""
    mapping = {}
    if os.path.exists(C.TOPSCORERS_FILE):
        with open(C.TOPSCORERS_FILE, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                key = (row.get("participant") or "").strip().lower()
                val = (row.get("top_scorer") or "").strip()
                if key:
                    mapping[key] = val or None
    return mapping


def _find_topscorer_in_sheet(ws):
    """Scan for a 'Toppscorer'-style label and read the cell to its right."""
    for row in ws.iter_rows(max_row=C.SCAN_LAST_ROW):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower().rstrip(":") in C.TOPSCORER_LABELS:
                neighbour = ws.cell(row=cell.row, column=cell.column + 1).value
                if isinstance(neighbour, str) and neighbour.strip():
                    return neighbour.strip()
    return None


def parse_workbook(path, topscorer_csv):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[C.SHEET_NAME]
    name = _participant_name(path)

    # --- group match predicted scores -------------------------------------
    group_scores = {}
    for i in range(C.GROUP_MATCH_COUNT):
        row = C.SCHEDULE_FIRST_ROW + i
        no = _cell(ws, C.COL_MATCH_NO, row)
        hg = _cell(ws, C.COL_HOME_GOALS, row)
        ag = _cell(ws, C.COL_AWAY_GOALS, row)
        if no is None:
            continue
        if isinstance(hg, (int, float)) and isinstance(ag, (int, float)):
            group_scores[str(int(no))] = [int(hg), int(ag)]

    # --- group winners / runners-up ----------------------------------------
    group_winners, group_runners = {}, {}
    for b in range(C.GROUP_BLOCK_COUNT):
        first = C.GROUP_BLOCK_FIRST_ROW + b * C.GROUP_BLOCK_STEP
        label = _cell(ws, C.COL_GROUP_LABEL, first)
        if not (label and str(label).lower().startswith("group")):
            continue
        label = str(label)
        for r in range(first, first + 4):
            place = _cell(ws, C.COL_PLACE, r)
            team = C.canonical_team(_cell(ws, C.COL_PLACE_TEAM, r))
            if place == 1 and team:
                group_winners[label] = team
            elif place == 2 and team:
                group_runners[label] = team

    # --- knockout bracket (cumulative team sets) ---------------------------
    bracket = {}
    for rnd, col in C.KO_TEAM_COLUMNS.items():
        teams = []
        for r in range(C.SCHEDULE_FIRST_ROW, C.SCAN_LAST_ROW + 1):
            t = C.canonical_team(_cell(ws, col, r))
            if t and t not in teams:
                teams.append(t)
        bracket[rnd] = teams

    # Final: take the two teams under the "Final" header, stopping at the
    # "Third-Place Play-Off" header.
    finalists = []
    in_final = False
    for r in range(C.SCHEDULE_FIRST_ROW, C.SCAN_LAST_ROW + 1):
        hdr = _cell(ws, C.COL_FINAL_HEADER, r)
        if isinstance(hdr, str):
            if hdr.strip() == C.FINAL_HEADER_TEXT:
                in_final = True
                continue
            if hdr.strip() == C.THIRD_PLACE_HEADER_TEXT:
                in_final = False
        if in_final:
            t = C.canonical_team(_cell(ws, C.COL_FINAL_TEAM, r))
            if t and t not in finalists:
                finalists.append(t)
    bracket["Final"] = finalists[:2]

    # --- champion ----------------------------------------------------------
    champion = C.canonical_team(_cell(ws, "CK", 68))

    # --- top scorer --------------------------------------------------------
    top_scorer = _find_topscorer_in_sheet(ws) or topscorer_csv.get(name.lower())

    return {
        "participant": name,
        "source_file": os.path.basename(path),
        "group_scores": group_scores,
        "group_winners": group_winners,
        "group_runners": group_runners,
        "bracket": bracket,
        "champion": champion,
        "top_scorer": top_scorer,
    }


def main():
    os.makedirs(C.PREDICTIONS_DIR, exist_ok=True)
    topscorer_csv = _load_topscorer_csv()
    paths = sorted(glob.glob(os.path.join(C.PARTICIPANTS_DIR, "*.xlsx")))
    paths = [p for p in paths if not os.path.basename(p).startswith("~$")]
    if not paths:
        print("No participant workbooks found in", C.PARTICIPANTS_DIR)
        return
    for path in paths:
        pred = parse_workbook(path, topscorer_csv)
        out = os.path.join(C.PREDICTIONS_DIR, f"{pred['participant']}.json")
        with open(out, "w", encoding="utf-8") as f:
            json.dump(pred, f, ensure_ascii=False, indent=2)
        ts = pred["top_scorer"] or "—"
        print(f"  {pred['participant']:<20} "
              f"{len(pred['group_scores'])} scores, champion={pred['champion']}, top scorer={ts}")
    print(f"Parsed {len(paths)} participant(s) into {C.PREDICTIONS_DIR}")


if __name__ == "__main__":
    main()
